# Test of rag evaluation
# Imports
import sys
from pprint import pprint

import openpyxl

sys.path.append("./dev/")
from components import DatasetHelpers, RagPipe, VectorStore

# Define API ENDPOINTS
LLM_URL = "http://10.103.251.104:8040/v1"
LLM_NAME = "llama3"
MARQO_URL = "http://10.103.251.104:8882"

# Load QM queries
datasetHelpers = DatasetHelpers()
_, queries, _ = datasetHelpers.loadQM()


# Load the VectorStore
documentDB = VectorStore(MARQO_URL)  # Connect to marqo client via python API
print(documentDB.getIndexes())  # Print all indexes
documentDB.connectIndex("ait-qm")  # Connect to the miniwikiindex


# Load the RagPipe
pipe = RagPipe()
pipe.connectVectorStore(documentDB)
pipe.connectLLM(LLM_URL, LLM_NAME)


# Run rag pipeline for test quieries of qm dataset
pipe.run(queries, None, None, newIngest=False, maxDocs=0, maxQueries=1)
for rag_element in pipe.rag_elements:
    print(rag_element["question"])
    print(rag_element["answer"])
    pprint(rag_element["contexts"])
    print("\n\n")


# Define the header for vairable number of contexts
header = ["Question", "Answer"]
n_contexts = len(pipe.rag_elements[0]["contexts"])
print(f"Number of contexts: {n_contexts}")
for i in range(1, n_contexts + 1):
    header.append(f"Context {i}")
print(f"Header: {header}")
# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the header to the first row
for col_num, header_title in enumerate(header, 1):
    sheet.cell(row=1, column=col_num, value=header_title)

# Add data dynamically
for row_num, rag_element in enumerate(pipe.rag_elements, 2):
    row_data = [rag_element["question"], rag_element["answer"]]
    for context in rag_element["contexts"]:
        row_data.append(context)

    for col_num, cell_value in enumerate(row_data, 1):
        sheet.cell(row=row_num, column=col_num, value=cell_value)

# Save the workbook to a file
file_path = "./data/QM_QA_noPrePostContext_noReRank.xlsx"
workbook.save(file_path)

print(f"Data written to {file_path}")


# Run the rag pipeline and ingest
# answer, contexts, ids = pipe.answerQuery(
#     "Whom do I have to talk to when going on a business trip?"
# )
# pprint(answer)
